import uuid
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import CurrentUser, get_db, require_roles
from app.models.plataforma import Proveedor
from app.schemas.proveedores import ImportacionExcelOut, MuestraAQLOut, ProveedorCreate, ProveedorOut
from app.services.aql import AQLError, calcular_plan_muestreo

router = APIRouter(prefix="/proveedores", tags=["proveedores"])


@router.post("", response_model=ProveedorOut, status_code=status.HTTP_201_CREATED)
def crear_proveedor(
    payload: ProveedorCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(require_roles("supervisor", "gerente")),
):
    proveedor = Proveedor(empresa_id=current_user.empresa_id, nombre=payload.nombre, nit=payload.nit)
    db.add(proveedor)
    db.commit()
    db.refresh(proveedor)
    return proveedor


@router.get("", response_model=list[ProveedorOut])
def listar_proveedores(
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(require_roles("supervisor", "gerente", "auditor", "operario")),
):
    return db.scalars(select(Proveedor).order_by(Proveedor.nombre)).all()


@router.get("/{proveedor_id}", response_model=ProveedorOut)
def consultar_proveedor(
    proveedor_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(require_roles("supervisor", "gerente", "auditor", "operario")),
):
    proveedor = db.get(Proveedor, proveedor_id)
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    return proveedor


@router.post("/importar-excel", response_model=ImportacionExcelOut, status_code=status.HTTP_201_CREATED)
async def importar_proveedores_excel(
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(require_roles("gerente")),
):
    """Carga masiva de proveedores desde un Excel. Solo el gerente puede
    hacerlo. El archivo debe tener una fila de encabezado con al menos una
    columna 'nombre' (obligatoria) y opcionalmente 'nit' -- el orden de las
    columnas no importa, se buscan por nombre de encabezado."""
    from openpyxl import load_workbook

    contenido = await archivo.read()
    try:
        libro = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="No se pudo leer el archivo. ¿Es un .xlsx válido?")

    hoja = libro.active
    filas = list(hoja.iter_rows(values_only=True))
    if not filas:
        raise HTTPException(status_code=422, detail="El archivo está vacío")

    encabezados = [str(c).strip().lower() if c else "" for c in filas[0]]
    if "nombre" not in encabezados:
        raise HTTPException(status_code=422, detail="El archivo debe tener una columna llamada 'nombre'")

    idx_nombre = encabezados.index("nombre")
    idx_nit = encabezados.index("nit") if "nit" in encabezados else None

    creados = 0
    omitidos = 0
    for fila in filas[1:]:
        if not fila or idx_nombre >= len(fila) or not fila[idx_nombre]:
            omitidos += 1
            continue
        nombre = str(fila[idx_nombre]).strip()
        nit = str(fila[idx_nit]).strip() if idx_nit is not None and idx_nit < len(fila) and fila[idx_nit] else None

        db.add(Proveedor(empresa_id=current_user.empresa_id, nombre=nombre, nit=nit))
        creados += 1

    db.commit()
    return ImportacionExcelOut(creados=creados, omitidos=omitidos)


@router.get("/{proveedor_id}/muestra-aql", response_model=MuestraAQLOut)
def calcular_muestra_aql(
    proveedor_id: uuid.UUID,
    tamano_lote: int = Query(..., ge=2),
    nivel_inspeccion_general: str = Query("II"),
    aql: float = Query(2.5),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(require_roles("supervisor", "gerente", "operario")),
):
    proveedor = db.get(Proveedor, proveedor_id)
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    try:
        plan = calcular_plan_muestreo(tamano_lote, nivel_inspeccion_general, aql)
    except AQLError as e:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(e))

    return MuestraAQLOut(
        severidad_actual=proveedor.nivel_inspeccion_actual,
        aql=aql,
        codigo_letra=plan.codigo_letra,
        tamano_muestra=plan.tamano_muestra,
        limite_aceptacion=plan.limite_aceptacion,
        limite_rechazo=plan.limite_rechazo,
    )